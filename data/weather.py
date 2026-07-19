# data/weather.py
"""
Módulo para ingesta de datos climáticos: clima histórico local y pronóstico a 7 días.
"""

import requests
import datetime
import pandas as pd
import numpy as np

# Coordenadas aproximadas para San Martín / Urdinarrain (Entre Ríos, Argentina)
LATITUDE = -32.687
LONGITUDE = -58.885

def get_7day_forecast(lat: float = LATITUDE, lon: float = LONGITUDE) -> list:
    """
    Obtiene el pronóstico de 7 días (lluvia y ETP aproximada) usando la API de Open-Meteo.
    Retorna una lista de diccionarios con la fecha, lluvia proyectada (mm) y ETP proyectada (mm).
    """
    url = f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lon}&daily=precipitation_sum,et0_fao_evapotranspiration&timezone=America/Argentina/Buenos_Aires"
    
    try:
        response = requests.get(url, timeout=5)
        if response.status_code == 200:
            data = response.json()
            daily = data.get("daily", {})
            dates = daily.get("time", [])
            precip = daily.get("precipitation_sum", [])
            et0 = daily.get("et0_fao_evapotranspiration", [])
            
            forecast = []
            for i in range(len(dates)):
                f_date = datetime.datetime.strptime(dates[i], "%Y-%m-%d").date()
                forecast.append({
                    "fecha": f_date,
                    "lluvia": float(precip[i]) if precip[i] is not None else 0.0,
                    "etp": float(et0[i]) if et0[i] is not None else 4.0 # default 4mm ETP
                })
            return forecast
    except Exception as e:
        print(f"Error consultando el pronóstico en Open-Meteo: {e}. Usando fallback...")
    
    # Fallback si falla la API
    today = datetime.date.today()
    forecast = []
    # Simular una serie de 7 días con clima veraniego promedio para Entre Ríos
    # Lluvias esporádicas y ETP promedio de 5mm en diciembre/enero
    for i in range(1, 8):
        f_date = today + datetime.timedelta(days=i)
        # 10% de probabilidad de lluvia
        lluvia_sim = 15.0 if i == 3 else 0.0 # lluvia simulada el día 3
        forecast.append({
            "fecha": f_date,
            "lluvia": lluvia_sim,
            "etp": 5.2 # promedio verano
        })
    return forecast


import os

class WeatherHistory:
    def __init__(self, excel_path: str = None):
        if excel_path is None:
            # El Excel de referencia vive en data/ junto a este módulo
            base_dir = os.path.dirname(os.path.abspath(__file__))
            self.excel_path = os.path.join(base_dir, "Balance hidrico - Maiz - El Trebol.xlsx")
        else:
            self.excel_path = excel_path
        self._cache_data = None

    def load_from_excel(self) -> pd.DataFrame:
        """Carga y procesa los datos históricos de clima de la planilla Excel de referencia."""
        if self._cache_data is not None:
            return self._cache_data
            
        import openpyxl
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        sheet = wb["ET - SGR"]
        
        records = []
        for r in range(2, 154):
            date_val = sheet.cell(r, 1).value
            if date_val is None or str(date_val).strip() == "Total":
                break
            
            # Si el valor de la fecha viene como datetime, lo convertimos
            if isinstance(date_val, datetime.datetime):
                fecha = date_val.date()
            else:
                fecha = pd.to_datetime(date_val).date()
                
            lluvia = sheet.cell(r, 4).value
            coef = sheet.cell(r, 5).value
            etp = sheet.cell(r, 8).value
            riego = sheet.cell(r, 3).value
            
            records.append({
                "fecha": fecha,
                "lluvia": float(lluvia) if lluvia is not None else 0.0,
                "coef_lluvia": float(coef) if coef is not None else 1.0,
                "etp": float(etp) if etp is not None else 0.0,
                "riego": float(riego) if riego is not None else 0.0
            })
            
        df = pd.DataFrame(records)
        self._cache_data = df
        return df


def _etp_estacional(month: int) -> float:
    """ETo promedio estacional (mm/día) para Entre Ríos según el mes."""
    return 5.5 if month in [11, 12, 1, 2, 3] else 2.0


def get_demo_forecast(df_history: pd.DataFrame, eval_date: datetime.date, days: int = 7) -> list:
    """
    Construye el "pronóstico" del modo demo a partir del propio histórico de la campaña
    (pronóstico perfecto): toma lluvia y ETP de los días posteriores a eval_date en la
    planilla. Si la campaña termina antes de completar la ventana, rellena con el
    promedio estacional. Evita mezclar el pronóstico real de hoy (otra estación del año)
    con fechas simuladas de la campaña.
    """
    df_future = df_history[df_history["fecha"] > eval_date].head(days)
    by_date = {row["fecha"]: row for _, row in df_future.iterrows()}

    forecast = []
    for i in range(1, days + 1):
        f_date = eval_date + datetime.timedelta(days=i)
        if f_date in by_date:
            row = by_date[f_date]
            forecast.append({"fecha": f_date, "lluvia": row["lluvia"], "etp": row["etp"]})
        else:
            forecast.append({"fecha": f_date, "lluvia": 0.0, "etp": _etp_estacional(f_date.month)})
    return forecast


def get_production_history(siembra_date: datetime.date, lat: float = None, lon: float = None) -> tuple:
    """
    Construye la serie climática diaria real desde la siembra hasta ayer para alimentar
    el motor en modo producción: base Open-Meteo Archive, sobrescrita por las estaciones
    DGHyOS (Urdinarrain, y Concepción del Uruguay como fallback) en los días que reportan
    datos sanos. El riego aplicado se inicializa en 0 (la carga de riegos reales queda
    pendiente de ingesta).

    Retorna una tupla (DataFrame con columnas fecha/lluvia/coef_lluvia/etp/riego, origen_str).
    """
    lat = lat if lat is not None else LATITUDE
    lon = lon if lon is not None else LONGITUDE
    end_date = datetime.date.today() - datetime.timedelta(days=1)
    if siembra_date > end_date:
        return pd.DataFrame(columns=["fecha", "lluvia", "coef_lluvia", "etp", "riego"]), "Sin datos"

    # 1. Base: Open-Meteo Archive para toda la temporada
    base_data = {}
    url_meteo = (
        f"https://archive-api.open-meteo.com/v1/archive?latitude={lat}&longitude={lon}"
        f"&start_date={siembra_date.strftime('%Y-%m-%d')}&end_date={end_date.strftime('%Y-%m-%d')}"
        f"&daily=precipitation_sum,et0_fao_evapotranspiration&timezone=America/Argentina/Buenos_Aires"
    )
    try:
        response = requests.get(url_meteo, timeout=10)
        if response.status_code == 200:
            daily = response.json().get("daily", {})
            dates = daily.get("time", [])
            precip = daily.get("precipitation_sum", [])
            et0 = daily.get("et0_fao_evapotranspiration", [])
            for i in range(len(dates)):
                d_obj = datetime.datetime.strptime(dates[i], "%Y-%m-%d").date()
                base_data[d_obj] = {
                    "lluvia": float(precip[i]) if precip[i] is not None else 0.0,
                    "etp": float(et0[i]) if et0[i] is not None else _etp_estacional(d_obj.month),
                    "origen": "Open-Meteo"
                }
    except Exception as e:
        print(f"Error cargando Open-Meteo Archive para la temporada: {e}")

    # 2. Sobrescribir con estaciones DGHyOS (validando sanidad estacional de ETo)
    station_days = 0
    for station_url, station_name in [
        ("https://www.hidraulica.gob.ar/ema/ema-urdinarrain/downld08.txt", "Urdinarrain (DGHyOS)"),
        ("https://www.hidraulica.gob.ar/ema/ema-curuguay/downld08.txt", "Concepción del Uruguay (DGHyOS)"),
    ]:
        station_data = fetch_and_parse_dg_hy_os(station_url)
        if not station_data:
            continue
        avg_et = sum(v["et"] for v in station_data.values()) / len(station_data)
        current_month = datetime.date.today().month
        min_et, max_et = (1.0, 3.5) if current_month in [4, 5, 6, 7, 8, 9] else (4.0, 9.0)
        if not (min_et <= avg_et <= max_et):
            print(f"{station_name}: ETo promedio ({avg_et:.2f} mm/d) fuera de rango estacional. Descartada.")
            continue
        for d_obj, vals in station_data.items():
            if siembra_date <= d_obj <= end_date and base_data.get(d_obj, {}).get("origen") != "DGHyOS":
                base_data[d_obj] = {"lluvia": vals["rain"], "etp": vals["et"], "origen": "DGHyOS"}
                station_days += 1

    if not base_data:
        return pd.DataFrame(columns=["fecha", "lluvia", "coef_lluvia", "etp", "riego"]), "Sin datos"

    # 3. Serie continua desde siembra hasta ayer (huecos -> promedio estacional)
    records = []
    seasonal_days = 0
    curr = siembra_date
    while curr <= end_date:
        if curr in base_data:
            day = base_data[curr]
        else:
            day = {"lluvia": 0.0, "etp": _etp_estacional(curr.month), "origen": "Estacional"}
            seasonal_days += 1
        records.append({
            "fecha": curr,
            "lluvia": day["lluvia"],
            "coef_lluvia": None,  # sin override: se usan los rangos de lluvia efectiva
            "etp": day["etp"],
            "riego": 0.0
        })
        curr += datetime.timedelta(days=1)

    df = pd.DataFrame(records)
    open_meteo_days = len(df) - station_days - seasonal_days
    parts = []
    if station_days > 0:
        parts.append(f"DGHyOS ({station_days}d)")
    if open_meteo_days > 0:
        parts.append(f"Open-Meteo Archive ({open_meteo_days}d)")
    if seasonal_days > 0:
        parts.append(f"Estimado estacional ({seasonal_days}d)")
    return df, " + ".join(parts) if parts else "Sin datos"


def fetch_and_parse_dg_hy_os(url: str) -> dict:
    """
    Descarga y parsea el archivo de reporte NOAA en formato texto de la DGHyOS.
    Retorna un diccionario de {date_obj: {'rain': float, 'et': float}} o None si falla.
    """
    try:
        response = requests.get(url, timeout=10, headers={'User-Agent': 'Mozilla/5.0'})
        if response.status_code != 200:
            print(f"Error {response.status_code} al descargar reporte NOAA de: {url}")
            return None
        content = response.text
    except Exception as e:
        print(f"Excepción al descargar reporte NOAA de {url}: {e}")
        return None

    lines = content.splitlines()
    separator_idx = -1
    for idx, line in enumerate(lines):
        if line.strip().startswith('----') and len(line.strip()) > 50:
            separator_idx = idx
            break
            
    if separator_idx == -1:
        print(f"No se encontró línea divisoria en el reporte de: {url}")
        return None
        
    header_line = lines[separator_idx - 1]
    
    # Mapear dinámicamente las columnas para tolerar diferencias de orden/nombres
    import re
    tokens = []
    for match in re.finditer(r'\S+', header_line):
        tokens.append({
            'name': match.group(),
            'start': match.start(),
            'end': match.end()
        })
        
    date_col_idx = -1
    rain_col_idx = -1
    et_col_idx = -1
    
    for idx, t in enumerate(tokens):
        name_lower = t['name'].lower()
        if name_lower in ['fecha', 'date']:
            date_col_idx = idx
        elif name_lower == 'rain':
            rain_col_idx = idx
        elif name_lower == 'et':
            et_col_idx = idx
            
    if date_col_idx == -1 or rain_col_idx == -1 or et_col_idx == -1:
        print(f"Columnas requeridas no encontradas en {url}. Fecha/Date Col: {date_col_idx}, Rain Col: {rain_col_idx}, ET Col: {et_col_idx}")
        return None
        
    daily_data = {}
    
    for r in range(separator_idx + 1, len(lines)):
        line = lines[r]
        if not line.strip():
            continue
            
        if any(keyword in line.upper() for keyword in ["MONTHLY", "AVERAGE", "TOTAL", "MAX", "MIN", "MEAN"]):
            break
            
        row_tokens = line.split()
        if len(row_tokens) < max(date_col_idx, rain_col_idx, et_col_idx) + 1:
            continue
            
        date_str = row_tokens[date_col_idx]
        if not re.match(r'^\d{2}/\d{2}/\d{2}$', date_str):
            continue
            
        # Parsear fecha
        try:
            date_obj = datetime.datetime.strptime(date_str, "%d/%m/%y").date()
        except Exception:
            continue
            
        # Extraer lluvia y ET manejando guiones y valores no numéricos
        rain_str = row_tokens[rain_col_idx]
        et_str = row_tokens[et_col_idx]
        
        rain_val = 0.0
        if rain_str != '---':
            try:
                rain_val = float(rain_str)
            except ValueError:
                pass
                
        et_val = 0.0
        if et_str != '---':
            try:
                et_val = float(et_str)
            except ValueError:
                pass
                
        if date_obj not in daily_data:
            daily_data[date_obj] = {
                'rain': 0.0,
                'et': 0.0
            }
            
        daily_data[date_obj]['rain'] += rain_val
        daily_data[date_obj]['et'] += et_val
        
    return daily_data


def get_realtime_weather_last_7_days(lat: float = LATITUDE, lon: float = LONGITUDE) -> tuple:
    """
    Obtiene la lluvia y ET diaria para cada uno de los últimos 7 días.
    Realiza una búsqueda día por día: si la estación principal (Urdinarrain) no tiene datos
    para un día específico, intenta completarlo con Concepción del Uruguay, y si también
    falla, con Open-Meteo Archive o promedios estacionales.
    Retorna una tupla (lista_diccionarios, active_source_str).
    """
    urdinarrain_url = "https://www.hidraulica.gob.ar/ema/ema-urdinarrain/downld08.txt"
    curuguay_url = "https://www.hidraulica.gob.ar/ema/ema-curuguay/downld08.txt"
    
    print("Descargando reportes para reconciliación día por día...")
    urdinarrain_data = fetch_and_parse_dg_hy_os(urdinarrain_url)
    curuguay_data = fetch_and_parse_dg_hy_os(curuguay_url)
    
    current_month = datetime.date.today().month
    is_winter = current_month in [4, 5, 6, 7, 8, 9]
    min_et, max_et = (1.0, 3.5) if is_winter else (4.0, 9.0)
    
    # Validar sanidad de Urdinarrain como estación completa
    if urdinarrain_data:
        avg_et = sum(v['et'] for v in urdinarrain_data.values()) / len(urdinarrain_data)
        if not (min_et <= avg_et <= max_et):
            print(f"Urdinarrain ETo promedio ({avg_et:.2f} mm/d) fuera de rango estacional ({min_et}-{max_et} mm/d). Descartando estación completa.")
            urdinarrain_data = None
            
    # Validar sanidad de Concepción como estación completa
    if curuguay_data:
        avg_et = sum(v['et'] for v in curuguay_data.values()) / len(curuguay_data)
        if not (min_et <= avg_et <= max_et):
            print(f"Concepción ETo promedio ({avg_et:.2f} mm/d) fuera de rango estacional ({min_et}-{max_et} mm/d). Descartando estación completa.")
            curuguay_data = None

    # 2. Descargar datos históricos de Open-Meteo para rellenar huecos parciales
    today = datetime.date.today()
    start_date = today - datetime.timedelta(days=7)
    end_date = today - datetime.timedelta(days=1)
    
    archive_data = {}
    url_meteo = f"https://archive-api.open-meteo.com/v1/archive?latitude={lat}&longitude={lon}&start_date={start_date.strftime('%Y-%m-%d')}&end_date={end_date.strftime('%Y-%m-%d')}&daily=precipitation_sum,et0_fao_evapotranspiration&timezone=America/Argentina/Buenos_Aires"
    
    try:
        response = requests.get(url_meteo, timeout=5)
        if response.status_code == 200:
            api_data = response.json()
            daily = api_data.get("daily", {})
            dates = daily.get("time", [])
            precip = daily.get("precipitation_sum", [])
            et0 = daily.get("et0_fao_evapotranspiration", [])
            for i in range(len(dates)):
                d_obj = datetime.datetime.strptime(dates[i], "%Y-%m-%d").date()
                archive_data[d_obj] = {
                    "rain": float(precip[i]) if precip[i] is not None else 0.0,
                    "etp": float(et0[i]) if et0[i] is not None else 2.5
                }
    except Exception as e:
        print(f"Error cargando fallback Open-Meteo para días faltantes: {e}")

    # 3. Construir la serie de 7 días rellenando individualmente
    result = []
    urdinarrain_days = 0
    curuguay_days = 0
    archive_days = 0
    seasonal_days = 0
    
    for i in range(7, 0, -1):
        target_date = today - datetime.timedelta(days=i)
        
        # A. Intentar Urdinarrain
        if urdinarrain_data and target_date in urdinarrain_data:
            day_rain = urdinarrain_data[target_date]["rain"]
            day_et = urdinarrain_data[target_date]["et"]
            urdinarrain_days += 1
            source_tag = "Urdinarrain (DGHyOS)"
        # B. Intentar Concepción del Uruguay (Fallback Primario)
        elif curuguay_data and target_date in curuguay_data:
            day_rain = curuguay_data[target_date]["rain"]
            day_et = curuguay_data[target_date]["et"]
            curuguay_days += 1
            source_tag = "Concepción del Uruguay (DGHyOS)"
        # C. Intentar Open-Meteo (Fallback Secundario)
        elif target_date in archive_data:
            day_rain = archive_data[target_date]["rain"]
            day_et = archive_data[target_date]["etp"]
            archive_days += 1
            source_tag = "Open-Meteo (Satélite)"
        # D. Promedio estacional último recurso
        else:
            day_rain = 0.0
            day_et = 5.5 if today.month in [11, 12, 1, 2, 3] else 2.0
            seasonal_days += 1
            source_tag = "Estimado (Promedio Estacional)"
            
        result.append({
            "fecha": target_date,
            "lluvia": day_rain,
            "etp": day_et,
            "origen": source_tag
        })
        
    # Construir resumen del origen de datos para la UI
    summary_sources = []
    if urdinarrain_days > 0:
        summary_sources.append(f"Urdinarrain ({urdinarrain_days}d)")
    if curuguay_days > 0:
        summary_sources.append(f"Concepción ({curuguay_days}d)")
    if archive_days > 0:
        summary_sources.append(f"Open-Meteo ({archive_days}d)")
    if seasonal_days > 0:
        summary_sources.append(f"Estimado ({seasonal_days}d)")
        
    active_source = " + ".join(summary_sources) if summary_sources else "Sin datos"
    return result, active_source

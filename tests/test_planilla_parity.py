# tests/test_planilla_parity.py
"""
Test de paridad: el motor de balance + la tabla de Kc deben reproducir día a día
la columna "AU: Riego" de la planilla de referencia El Trébol 2025-26.
Es la validación de referencia antes de cargar los datos reales de San Martín.
"""

import datetime
import os
import openpyxl
import pytest

from core.balance import DailyWaterBalance
from core.crop import CropModel

EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "data", "Balance hidrico - Maiz - El Trebol.xlsx"
)
SIEMBRA = datetime.date(2025, 9, 17)


def _load_planilla_rows():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet = wb["ET - SGR"]
    rows = []
    for r in range(2, 154):
        d = sheet.cell(r, 1).value
        if d is None or str(d).strip() == "Total":
            break
        rows.append({
            "fecha": d.date() if isinstance(d, datetime.datetime) else d,
            "riego": float(sheet.cell(r, 3).value or 0),
            "lluvia": float(sheet.cell(r, 4).value or 0),
            "coef": float(sheet.cell(r, 5).value) if sheet.cell(r, 5).value is not None else None,
            "etp": float(sheet.cell(r, 8).value or 0),
            "kc": float(sheet.cell(r, 9).value or 0),
            "au": float(sheet.cell(r, 11).value) if sheet.cell(r, 11).value is not None else None,
        })
    return rows


@pytest.fixture(scope="module")
def planilla():
    return _load_planilla_rows()


def test_tabla_kc_replica_planilla(planilla):
    """La tabla KC_DAS_TABLE debe coincidir con la columna KC de la planilla."""
    crop = CropModel(SIEMBRA)
    for row in planilla:
        kc_model, _ = crop.get_kc_and_stage(row["fecha"])
        assert kc_model == pytest.approx(row["kc"], abs=1e-6), (
            f"Kc divergente el {row['fecha']}: modelo {kc_model} vs planilla {row['kc']}"
        )


def test_balance_replica_planilla(planilla):
    """El AU calculado día a día debe coincidir con la columna 'AU: Riego' (tolerancia 0.01 mm)."""
    engine = DailyWaterBalance(techo_sistema=120.0)
    au_prev = 90.0
    au_secano_prev = 90.0
    for row in planilla:
        res = engine.calculate_next_day(
            au_prev=au_prev,
            au_secano_prev=au_secano_prev,
            riego=row["riego"],
            lluvia=row["lluvia"],
            etp=row["etp"],
            kc=row["kc"],
            coef_override=row["coef"],
        )
        if row["au"] is not None:
            assert res["au_riego"] == pytest.approx(row["au"], abs=0.01), (
                f"AU divergente el {row['fecha']}: modelo {res['au_riego']} vs planilla {row['au']}"
            )
        au_prev = res["au_riego"]
        au_secano_prev = res["au_secano"]
